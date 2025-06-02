import os
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def clean_illegal_chars(text):
    if not isinstance(text, str):
        return text
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", text)

def extract_realtime_entries(file_path):
    """
    Extract all 'real time' entries from a log file.
    Returns a list of tuples: (Author, File Location, Filename, Timestamp, Log Snippet, Raw Value, Real Time (seconds))
    """
    entries = []
    lines = []
    file_timestamp = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
    
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()

    # Extract author
    author = ''
    for line in lines:
        if 'author' in line.lower():
            author_match = re.search(r'author\s*[:\-]\s*(.*)', line, re.IGNORECASE)
            if author_match:
                author = author_match.group(1).strip()
                break

    # Find all real time entries
    for i, line in enumerate(lines):
        if 'real time' in line.lower():
            line_lower = line.lower()
            raw_value = ''
            total_seconds = None

            # Format: real time 10:0.15
            colon_match = re.search(r"real time\s+([0-9]+):([0-9.]+)", line_lower)
            if colon_match:
                minutes = int(colon_match.group(1))
                seconds = float(colon_match.group(2))
                total_seconds = round(minutes * 60 + seconds, 2)
                raw_value = colon_match.group(0).split('real time')[-1].strip()
            else:
                # Format: real time 0.15 seconds or 1.2 minutes
                unit_match = re.search(r"real time\s+([0-9.]+)\s+(seconds|minutes)", line_lower)
                if unit_match:
                    time_val = float(unit_match.group(1))
                    unit = unit_match.group(2)
                    total_seconds = round(time_val * 60 if unit == 'minutes' else time_val, 2)
                    raw_value = unit_match.group(0).split('real time')[-1].strip()

            if total_seconds is not None:
                snippet_start = max(0, i - 20)
                for j in range(i - 1, snippet_start - 1, -1):
                    if 'real time' in lines[j].lower():
                        snippet_start = j + 1
                        break
                snippet = ''.join(lines[snippet_start:i + 1]).strip()
                entries.append((
                    author,
                    os.path.dirname(file_path),
                    os.path.basename(file_path),
                    file_timestamp,
                    snippet,
                    raw_value,
                    total_seconds
                ))
    return entries

def scan_logs(folder_path):
    summary_entries = []
    detail_entries = []
    
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".log"):
                file_path = os.path.join(root, file)
                all_entries = extract_realtime_entries(file_path)

                if all_entries:
                    # Sort entries by appearance (they are in order already)
                    last_entry = all_entries[-1]
                    summary_entries.append(last_entry)

                    # All but last with time > 5 sec
                    for entry in all_entries[:-1]:
                        if entry[-1] > 5:
                            detail_entries.append(entry)
                else:
                    # No real time entries, but still include in summary with blanks
                    file_timestamp = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                    author = ''
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            for line in f:
                                if 'author' in line.lower():
                                    match = re.search(r'author\s*[:\-]\s*(.*)', line, re.IGNORECASE)
                                    if match:
                                        author = match.group(1).strip()
                                        break
                    except Exception as e:
                        print(f"Error reading {file_path}: {e}")
                    summary_entries.append((author, root, file, file_timestamp, '', '', ''))
                    
    df_summary = pd.DataFrame(summary_entries, columns=["Author", "File Location", "Filename", "Timestamp", "Log Snippet", "Raw Value", "Real Time (seconds)"])
    df_details = pd.DataFrame(detail_entries, columns=["Author", "File Location", "Filename", "Timestamp", "Log Snippet", "Raw Value", "Real Time (seconds)"])
    
    return df_summary, df_details

def format_excel_output(file_path, wrap_col_name="Log Snippet"):
    wb = load_workbook(file_path)
    ws = wb["Details"]

    col_index = None
    for cell in ws[1]:
        if cell.value == wrap_col_name:
            col_index = cell.column
            break

    if col_index:
        col_letter = get_column_letter(col_index)
        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    num_lines = str(cell.value).count("\n") + 1
                    ws.row_dimensions[cell.row].height = max(15, min(15 * num_lines, 200))
        ws.column_dimensions[col_letter].width = 80

    wb.save(file_path)
    print(f"✨ Excel formatting applied to column '{wrap_col_name}'")

# === MAIN ===
folder_to_scan = r"J:\bdm\tbos\TAK279\studies\3001\dryrun1\programs"
timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
output_excel = fr"J:\bdm\tbos\TAK279\studies\3001\dryrun1\oversight\reports\real_time_report_{timestamp_str}.xlsx"

# Extract info
df_summary, df_details = scan_logs(folder_to_scan)

# Clean illegal characters in string fields
for df in [df_summary, df_details]:
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].apply(clean_illegal_chars)
        
# Sort by Real Time (seconds) descending
df_summary = df_summary.sort_values(by="Real Time (seconds)", ascending=False, na_position="last")

# Drop Log Snippet from Summary
df_summary = df_summary.drop(columns=["Log Snippet"], errors="ignore")

# Export to Excel
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    df_summary.to_excel(writer, index=False, sheet_name='Summary')
    df_details[["Author", "File Location", "Filename", "Log Snippet", "Raw Value", "Real Time (seconds)"]].to_excel(writer, index=False, sheet_name='Details')

format_excel_output(output_excel)

print(f"✅ Excel report created: {output_excel}")
